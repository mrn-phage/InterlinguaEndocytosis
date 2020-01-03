import re
import codecs

import pandas as pd
import openpyxl as opx


class DataFrameMaker():
    MODERN_EXCEL_PATTERN = (r'.*\.xlsx|.*\.xlsm', 'modern_excel')
    LEGACY_EXCEL_PATTERN = (r'.*\.xls', 'legacy_excel')
    CSV_PATTERN = (r'.*\.csv', 'csv')
    PATTERN_DICT = {'me': MODERN_EXCEL_PATTERN,
                    'le': LEGACY_EXCEL_PATTERN, 'csv': CSV_PATTERN}

    def __init__(self, file_path):
        """
        正規表現によるファイル種別の判定。
        判定に応じたデータフレーム
        """

        self.type = None
        self.sheet_name_list = []
        self.df_list = []
################################################################################
        # ファイル種別判定
        ext = re.fullmatch(DataFrameMaker.PATTERN_DICT['me'][0], file_path)
        if ext != None:
            self.type = DataFrameMaker.PATTERN_DICT['me'][1]

        ext = re.fullmatch(DataFrameMaker.PATTERN_DICT['le'][0], file_path)
        if ext != None:
            self.type = DataFrameMaker.PATTERN_DICT['le'][1]

        ext = re.fullmatch(DataFrameMaker.PATTERN_DICT['csv'][0], file_path)
        if ext != None:
            self.type = DataFrameMaker.PATTERN_DICT['csv'][1]

################################################################################
        # シート名が取得できるファイル種別ならシート名変数を生成
        if self.type == 'modern_excel':
            wb = opx.load_workbook(file_path)
            self.sheet_name_list = wb.sheetnames    # シート名を取得

        else:
            self.sheet_name_list = None

################################################################################
        # 正規表現によるファイル種別の判定と対応したデータフレームlistを生成

        # xlsx、或いはxlsmだったらこの処理でデータフレーム生成
        if self.type == 'modern_excel':
            wb = opx.load_workbook(file_path)

            # シート毎にデータフレームを生成
            for sheet_name in self.sheet_name_list:
                sheet = wb[sheet_name]
                data = sheet.values
                self.df_list.append(pd.DataFrame(data))

        # xlsだったらこの処理でデータフレーム生成
        if self.type == 'legacy_excel':
            pass

        # csvだったらこの処理でデータフレーム生成
        if self.type == 'csv':
            with codecs.open(file_path, "r", "cp932", "ignore") as f:
                self.df_list.append(pd.read_csv(f, header=None))
